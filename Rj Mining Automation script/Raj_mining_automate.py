from playwright.sync_api import sync_playwright
from playwright.sync_api import expect
import openpyxl as xl
import subprocess
import re
workbook = xl.load_workbook("RAJ_MINING_DEVICES.xlsx")
sheet = workbook["Sheet1"]
Devices=[]
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     imei , vehicle_no ,_= row
#     Devices.append((imei,vehicle_no))
# print(Devices)
Devices = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    imei, vehicle_no, _ = row
    if imei is None or vehicle_no is None:
        continue          # skip empty/deleted rows
    Devices.append((imei, vehicle_no))

print(Devices)
input("Enter after Confirming Imei and Vehicle number")
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    page.goto("https://minesportal.rajasthan.gov.in/")
    def ok():
        page.get_by_role("button", name="OK").click()
    def wait():
        page.wait_for_load_state("networkidle")
    ok()
    Un = page.get_by_placeholder("Username")
    Pw = page.get_by_placeholder("Password")
    Un.fill("VENDOR00219")
    Pw.fill("Watsoo@0909")
    page.wait_for_timeout(4000)
    def login():
        page.get_by_role("button", name="Login").click()
    login()
    ok()
    gps_vendor = page.get_by_role("heading",name="GPS Vendor")
    gps_vendor.click()
    device_tagging = page.get_by_text("Device Tagging")
    device_tagging.click()
    Reg_device = page.get_by_text("Register GPS Device")
    Reg_device.click()
    # imei = input("Enter IMEI number:")
    # search_imei = page.locator("input[placeholder='IMEI/Serial || Vehicle Regd.No']")
    # def select_when_ready(select_locator, label, timeout=10000):
                    # option = select_locator.locator("option").filter(
                    # has_text=re.compile(f"^{re.escape(label.strip())}$"))
                    # expect(option).to_be_attached(timeout=timeout)
                    # select_locator.select_option(label=label)
    for imei, vehicle_no in Devices:          
        imei = str(imei)                     
        vehicle_no = str(vehicle_no)

        search_imei = page.get_by_placeholder("IMEI / Serial || Vehicle Regd. No")
        print("its found" ,search_imei.count())
        search_imei.fill(imei)
        search = page.get_by_role("button", name="Search")
        search.click()
        no_data = page.get_by_text("NO DATA FOUND!!",exact=True)
        # no_data.wait_for_load_state("networkidle")
    
        if no_data.is_visible():
            print("IMEI is not registered")
            # add_devc(imei)
            add_dvc = page.get_by_role("button", name="Add Device")
            add_dvc.click()
            wait()
            vendor = page.locator('select[formcontrolname="vendorId"]')
            # # vendor.select_option(value="WATSOO EXPRESS PRIVATE LIMITED")
            # vendor.select_option(index=1)
            # input("Vendor id request is selcted")
            vendor.select_option(value="224")
            print("Raw DOM value after select_option:", vendor.input_value())
            page.wait_for_timeout(500)
            print("Raw DOM value after short wait:", vendor.input_value())
            vendor.select_option(value="224")
            vendor.dispatch_event("change")
            vendor.dispatch_event("input")
            page.wait_for_timeout(300)
            print("Value after manual dispatch:", vendor.input_value())

            
            # add_dvc.click()
            # vendor = page.locator('select[formcontrolname="vendorId"]')
            # select_when_ready(vendor, "WATSOO EXPRESS PRIVATE LIMITED")
            # input("Vendor is selected")
            # options_now = vendor.locator("option").all_text_contents()
            # print(f"[{imei}] Vendor options before selecting:", options_now)
            # select_when_ready(vendor, "WATSOO EXPRESS PRIVATE LIMITED")

            test_req = page.locator('select[formcontrolname="requestNo"]')
            test_req.select_option(label="REQVEN-1-305")
            input("Test request is selcted")

            manuf_name = page.locator('select[formcontrolname="manufacturerName"]')
            manuf_name.select_option(label="WATSOO EXPRESS PRIVATE LIMITED")
            input("Manufacturer name is selected")
            Enter_imei = page.get_by_role("textbox", name="Enter IMEI or Serial Number")
            Enter_imei.fill(imei)
            # Vin = input("Enter Vehicle Number:")
            Veh_no = page.get_by_placeholder("Enter vehicle registration")
            Veh_no.fill(vehicle_no)
            Verify = page.get_by_role("button", name="Verify").click()
            page.wait_for_timeout(5000)
            already_exit = page.get_by_text("Vehicle number already exist.")
            # if already_exit.is_visible():
            # print("Vehicle number already exist")
            # close= page.get_by_title("Close")
            # close.click()
            # else:
            Date = page.locator('[formcontrolname="installDateTime"]').click()
            Wait_state = page.locator('.flatpickr-calendar.open').wait_for(state="visible")
            today_date = page.locator('.flatpickr-calendar.open .flatpickr-day.today').click()
            # Date.click()
            image = page.locator("input[type='file']").set_input_files(r"C:\Users\Admin\Desktop\Playwright Leaning\4G Device pic.png")
            Sme_zone = page.locator('select[formcontrolname="smeOfficeName"]')
            Sme_zone.select_option(label="AJMER")
            Me_district = page.locator('select[formcontrolname="meOfficeName"]')
            Me_district.select_option(label="AME GOTAN")
            page.wait_for_load_state("networkidle")
            checkbox = page.locator('[formcontrolname="healthDeclarationAccepted"]').check()
            Register_device = page.get_by_role("button", name="Register Device")
            Register_device.click()
            ok()
            device_tagging.click()
            Reg_device.click()
            # device_tagging.click()
            # Reg_device.click()
        
        else:
            print("imei is already registered")
    subprocess.run(["python","rj_Mining_dp.py"])
    print("data is pushed in rajasthan mining portal after registering device")
    input("Enter to close ")
    browser.close()
# 861729079599595,RJ26GA9041